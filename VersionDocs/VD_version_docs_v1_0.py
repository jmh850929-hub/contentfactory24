#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
VD_version_docs_v1_0.py

CAPS VersionDocs Module v1.0
----------------------------------------
역할:
- CAPS AutoSync 패치 이후 생성되는 버전 정보를 가지고
  JSON / Excel / PDF 의 3중 버전 문서를 자동으로 생성한다.

출력:
- output/json/version_xxx.json
- output/excel/version_history.xlsx
- output/pdf/version_report_YYYY-MM-DD.pdf

의존:
- AutoSync가 넘겨주는 version_info (dict)
- templates/version_json_template.json (없으면 자동 생성)
"""

import json
import datetime
from pathlib import Path
from openpyxl import Workbook, load_workbook
from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet


# -----------------------------------------------------
# 1) 기본 설정
# -----------------------------------------------------
MODULE_VERSION = "1.0.0"
MODULE_NAME = "VersionDocs"


class VersionDocs:
    def __init__(self, base_dir: Path = None):
        self.base_dir = base_dir or Path(__file__).resolve().parent

        # 디렉터리 정의
        self.templates_dir = self.base_dir / "templates"
        self.output_dir = self.base_dir / "output"
        self.json_output = self.output_dir / "json"
        self.xlsx_output = self.output_dir / "excel"
        self.pdf_output = self.output_dir / "pdf"

        # 파일 정의
        self.template_json_path = self.templates_dir / "version_json_template.json"
        self.excel_path = self.xlsx_output / "version_history.xlsx"

        self._ensure_directories()
        self._ensure_template()

    # -----------------------------------------------------
    # 폴더 생성
    # -----------------------------------------------------
    def _ensure_directories(self):
        self.templates_dir.mkdir(parents=True, exist_ok=True)
        self.json_output.mkdir(parents=True, exist_ok=True)
        self.xlsx_output.mkdir(parents=True, exist_ok=True)
        self.pdf_output.mkdir(parents=True, exist_ok=True)

    # -----------------------------------------------------
    # 템플릿 자동 생성
    # -----------------------------------------------------
    def _ensure_template(self):
        if not self.template_json_path.exists():
            template_data = {
                "version": "",
                "module": "",
                "prev_version": "",
                "new_version": "",
                "date": "",
                "features_changed": [],
                "lts_applied": False,
                "notes": ""
            }
            with open(self.template_json_path, "w", encoding="utf-8") as f:
                json.dump(template_data, f, indent=2, ensure_ascii=False)

    # -----------------------------------------------------
    # JSON 생성
    # -----------------------------------------------------
    def create_json(self, version_info: dict):
        version = version_info.get("new_version", "unknown")
        output_file = self.json_output / f"version_{version}.json"

        with open(self.template_json_path, "r", encoding="utf-8") as f:
            template = json.load(f)

        template.update(version_info)

        with open(output_file, "w", encoding="utf-8") as f:
            json.dump(template, f, indent=2, ensure_ascii=False)

        return output_file

    # -----------------------------------------------------
    # Excel append
    # -----------------------------------------------------
    def create_excel(self, version_info: dict):
        if self.excel_path.exists():
            wb = load_workbook(self.excel_path)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append([
                "Date", "Module", "PrevVersion", "NewVersion",
                "Features", "LTS Applied", "Notes"
            ])

        ws.append([
            version_info.get("date", ""),
            version_info.get("module", MODULE_NAME),
            version_info.get("prev_version", ""),
            version_info.get("new_version", ""),
            ", ".join(version_info.get("features_changed", [])),
            version_info.get("lts_applied", False),
            version_info.get("notes", "")
        ])

        wb.save(self.excel_path)
        return self.excel_path

    # -----------------------------------------------------
    # PDF 생성
    # -----------------------------------------------------
    def create_pdf(self, version_info: dict):
        date = version_info.get("date", "")
        module = version_info.get("module", MODULE_NAME)
        new_ver = version_info.get("new_version", "unknown")

        output_file = self.pdf_output / f"version_report_{date}.pdf"

        styles = getSampleStyleSheet()
        story = []
        story.append(Paragraph(f"<b>CAPS Version Report</b>", styles['Title']))
        story.append(Paragraph(f"Module: {module}", styles['Normal']))
        story.append(Paragraph(f"Date: {date}", styles['Normal']))
        story.append(Paragraph(f"Version: {new_ver}", styles['Normal']))
        story.append(Paragraph("<br/>", styles['Normal']))
        story.append(Paragraph("<b>Features Changed:</b>", styles['Heading2']))

        for feat in version_info.get("features_changed", []):
            story.append(Paragraph(f"- {feat}", styles['Normal']))

        story.append(Paragraph("<br/>", styles['Normal']))
        story.append(Paragraph(f"LTS Applied: {version_info.get('lts_applied', False)}", styles['Normal']))
        story.append(Paragraph(f"Notes: {version_info.get('notes', '')}", styles['Normal']))

        pdf = SimpleDocTemplate(str(output_file))
        pdf.build(story)

        return output_file

    # -----------------------------------------------------
    # 메인 실행
    # -----------------------------------------------------
    def generate_docs(self, version_info: dict):
        json_file = self.create_json(version_info)
        xlsx_file = self.create_excel(version_info)
        pdf_file = self.create_pdf(version_info)

        return json_file, xlsx_file, pdf_file


# -----------------------------------------------------
# 엔트리 포인트
# -----------------------------------------------------
def main():
    docs = VersionDocs()

    # 테스트용 입력
    today = datetime.datetime.now().strftime("%Y-%m-%d")

    sample_version_info = {
        "module": MODULE_NAME,
        "prev_version": "0.0.0",
        "new_version": MODULE_VERSION,
        "date": today,
        "features_changed": ["Initial VersionDocs module added"],
        "lts_applied": True,
        "notes": "AutoSync 테스트용 기본 생성"
    }

    json_f, xlsx_f, pdf_f = docs.generate_docs(sample_version_info)

    print("[VersionDocs v1.0] 문서 생성 완료")
    print(f"JSON : {json_f}")
    print(f"Excel: {xlsx_f}")
    print(f"PDF  : {pdf_f}")


if __name__ == "__main__":
    main()
