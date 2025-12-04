# excel_writer.py
# AutoSync 4.0 - VersionDocs(Excel)

from openpyxl import Workbook
from datetime import datetime

def write_excel_version():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "VersionDoc Type"
    ws["B1"] = "Generated At"
    ws["A2"] = "excel_version_doc"
    ws["B2"] = datetime.now().isoformat()

    wb.save("AS_version_doc.xlsx")
    return True
